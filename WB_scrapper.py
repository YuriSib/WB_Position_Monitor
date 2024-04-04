import requests
import openpyxl


def get_searching_data():
    wb = openpyxl.load_workbook('Позиции по поисковым запросам.xlsx')
    ws = wb.active

    max_row = ws.max_row
    searching_data = {}
    for row in range(2, max_row+1):
        article = ws.cell(row=row, column=1).value
        qwery_list = ws.cell(row=row, column=2).value.split(', ')
        searching_data[article] = qwery_list

    return searching_data


class WBMonitor:
    def __init__(self, target_article):
        # self.id_list = id_list
        # self.key = key
        self.target_article = target_article

    @staticmethod
    def settings(page, key):
        headers = {
            'Accept': '*/*',
            'Accept-Language': 'ru,en;q=0.9',
            'Connection': 'keep-alive',
            'Origin': 'https://www.wildberries.ru',
            'Referer': 'https://www.wildberries.ru/catalog/0/search.aspx?search=%D0%B1%D0%B0%D1%88%D0%BC%D0%B0%D0%BA%D0%B8',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'cross-site',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "YaBrowser";v="24.1", "Yowser";v="2.5"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'x-queryid': 'qid481473304171208404820240402185417',
        }

        params = {
            'page': f'{page}',
            # 'ab_testing': 'false',
            'appType': '1',
            'curr': 'rub',
            'dest': '-1257786',
            'query': f'{key}',
            'resultset': 'catalog',
            'sort': 'popular',
            'suppressSpellcheck': 'false',
            # 'limit': '300'
        }
        response = requests.get('https://search.wb.ru/exactmatch/ru/common/v5/search', params=params,
                                headers=headers, timeout=60)
        if response.status_code != 200:
            return 0
        return response.json()

    @staticmethod
    def find_position(response, target_article, cnt=1):
        products_raw = response.get('data', {}).get('products', None)
        if products_raw != None and len(products_raw) > 0:
            for product in products_raw:
                # name = product.get('name', None)
                article = product.get('id', None)
                if article == target_article:
                    print(product.get('name', None))
                    return cnt
                cnt += 1
            return False

    def hoarder(self, qwery_key):
        page_num = 1
        position_num = 1
        finish_flag = False
        while True:
            response = self.settings(page=page_num, key=qwery_key)

            quantity_cards = len(response['data']['products'])
            if quantity_cards == 1:
                continue
            if quantity_cards != 100:
                finish_flag = True

            position = self.find_position(response=response, target_article=self.target_article, cnt=position_num) if response else []
            if position:
                return position
            else:
                if finish_flag or (position_num+100) > 1000:
                    return 'Не удалось установить позицию'
                page_num += 1
                position_num += 100
                print(f'Перебрано {position_num-1} карточек')


if __name__ == "__main__":
    key = 'футболка женская'
    target_article = 206854348
    wbm = WBMonitor(target_article=target_article)
    num_position = wbm.hoarder(key)
    print(f'Позиция артикула {target_article} по запросу "{key}" - {num_position}')
    # print(get_searching_data())
